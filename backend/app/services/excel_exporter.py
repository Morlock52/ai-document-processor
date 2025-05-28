import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import io

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Export extracted document data to formatted Excel files"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    async def export_to_excel(self, 
                            data: List[Dict[str, Any]], 
                            template: Optional[str] = None,
                            include_metadata: bool = True) -> bytes:
        """Export data to Excel with formatting and optional template"""
        try:
            # Create workbook
            wb = Workbook()
            
            # Main data sheet
            ws_data = wb.active
            ws_data.title = "Extracted Data"
            
            # Write data
            self._write_data_sheet(ws_data, data)
            
            # Add metadata sheet if requested
            if include_metadata:
                ws_meta = wb.create_sheet("Metadata")
                self._write_metadata_sheet(ws_meta, data)
            
            # Add summary sheet
            ws_summary = wb.create_sheet("Summary", 0)
            self._write_summary_sheet(ws_summary, data)
            
            # Apply template if provided
            if template:
                self._apply_template(wb, template)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.read()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _write_data_sheet(self, ws, data: List[Dict[str, Any]]):
        """Write main data to worksheet"""
        if not data:
            ws.append(["No data to export"])
            return
            
        # Get all unique fields
        all_fields = set()
        for record in data:
            all_fields.update(record.keys())
        
        # Sort fields for consistent ordering
        fields = sorted(list(all_fields))
        
        # Write headers
        headers = ["Document #"] + fields
        ws.append(headers)
        
        # Apply header formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            # Adjust column width
            ws.column_dimensions[get_column_letter(col)].width = max(15, len(header) + 2)
        
        # Write data rows
        for idx, record in enumerate(data, 1):
            row_data = [idx]
            for field in fields:
                value = record.get(field, "N/A")
                # Handle nested data
                if isinstance(value, (dict, list)):
                    value = str(value)
                row_data.append(value)
                
            ws.append(row_data)
            
            # Apply cell formatting
            for col in range(1, len(row_data) + 1):
                cell = ws.cell(row=idx + 1, column=col)
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _write_metadata_sheet(self, ws, data: List[Dict[str, Any]]):
        """Write metadata about the extraction process"""
        ws.append(["Extraction Metadata"])
        ws.append([])
        
        # Summary statistics
        ws.append(["Total Documents", len(data)])
        ws.append(["Extraction Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["AI Model Used", "GPT-4o"])
        ws.append([])
        
        # Field statistics
        ws.append(["Field Statistics"])
        ws.append(["Field Name", "Filled Count", "N/A Count", "Fill Rate %"])
        
        # Calculate field statistics
        field_stats = {}
        for record in data:
            for field, value in record.items():
                if field not in field_stats:
                    field_stats[field] = {"filled": 0, "na": 0}
                
                if value and value != "N/A":
                    field_stats[field]["filled"] += 1
                else:
                    field_stats[field]["na"] += 1
        
        # Write field statistics
        for field, stats in sorted(field_stats.items()):
            total = stats["filled"] + stats["na"]
            fill_rate = (stats["filled"] / total * 100) if total > 0 else 0
            ws.append([field, stats["filled"], stats["na"], f"{fill_rate:.1f}%"])
        
        # Format metadata sheet
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.border
                if cell.row == 1 or cell.row == 7:
                    cell.font = Font(bold=True, size=14)
                elif cell.row == 8:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
    
    def _write_summary_sheet(self, ws, data: List[Dict[str, Any]]):
        """Write summary dashboard"""
        ws.append(["Document Processing Summary"])
        ws['A1'].font = Font(bold=True, size=16)
        ws.append([])
        
        # Key metrics
        ws.append(["Key Metrics"])
        ws['A3'].font = Font(bold=True, size=14)
        
        metrics = [
            ["Total Documents Processed", len(data)],
            ["Average Fields per Document", self._calculate_avg_fields(data)],
            ["Data Completeness", f"{self._calculate_completeness(data):.1f}%"],
            ["Processing Date", datetime.now().strftime("%Y-%m-%d")]
        ]
        
        for metric in metrics:
            ws.append(metric)
        
        # Add chart if enough data
        if len(data) >= 5:
            self._add_completeness_chart(ws, data)
        
        # Format summary
        for row in ws.iter_rows(min_row=3, max_row=7):
            row[0].font = Font(bold=True)
            for cell in row:
                cell.border = self.border
    
    def _calculate_avg_fields(self, data: List[Dict[str, Any]]) -> float:
        """Calculate average number of fields per document"""
        if not data:
            return 0
        
        total_fields = sum(len(record) for record in data)
        return round(total_fields / len(data), 1)
    
    def _calculate_completeness(self, data: List[Dict[str, Any]]) -> float:
        """Calculate data completeness percentage"""
        if not data:
            return 0
        
        total_fields = 0
        filled_fields = 0
        
        for record in data:
            for value in record.values():
                total_fields += 1
                if value and value != "N/A":
                    filled_fields += 1
        
        return (filled_fields / total_fields * 100) if total_fields > 0 else 0
    
    def _add_completeness_chart(self, ws, data: List[Dict[str, Any]]):
        """Add a chart showing data completeness"""
        # Prepare chart data
        ws.append([])
        ws.append(["Field Completeness Analysis"])
        ws.append(["Field", "Completeness %"])
        
        field_completeness = {}
        for record in data:
            for field, value in record.items():
                if field not in field_completeness:
                    field_completeness[field] = {"filled": 0, "total": 0}
                
                field_completeness[field]["total"] += 1
                if value and value != "N/A":
                    field_completeness[field]["filled"] += 1
        
        # Write chart data
        chart_start_row = ws.max_row + 1
        for field, stats in sorted(field_completeness.items())[:10]:  # Top 10 fields
            completeness = (stats["filled"] / stats["total"] * 100) if stats["total"] > 0 else 0
            ws.append([field, completeness])
        
        # Create chart
        chart = BarChart()
        chart.title = "Field Completeness (%)"
        chart.x_axis.title = "Fields"
        chart.y_axis.title = "Completeness %"
        
        data_ref = Reference(ws, min_col=2, min_row=chart_start_row, 
                           max_row=ws.max_row, max_col=2)
        cats = Reference(ws, min_col=1, min_row=chart_start_row + 1, 
                        max_row=ws.max_row)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "E10")
    
    def _apply_template(self, wb: Workbook, template_name: str):
        """Apply a predefined template to the workbook"""
        # This would load and apply specific templates
        # For now, just a placeholder
        logger.info(f"Applying template: {template_name}")
        
    
    async def export_batch(self, documents: List[Dict[str, Any]], format: str = "xlsx") -> bytes:
        """Export multiple documents as separate sheets or files"""
        if format == "xlsx":
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for idx, doc in enumerate(documents):
                ws = wb.create_sheet(f"Document_{idx + 1}")
                self._write_data_sheet(ws, [doc])
            
            # Add combined sheet
            ws_all = wb.create_sheet("All Documents", 0)
            self._write_data_sheet(ws_all, documents)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
        
        else:
            # Other formats (CSV, etc.) can be implemented here
            raise ValueError(f"Unsupported format: {format}")

    async def export_template(self, documents: List[Dict[str, Any]], field_names: List[str]) -> bytes:
        """Export documents in template format with detected fields as columns"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template Data"

            # Prepare header row with all detected fields
            headers = ["Source Document"] + field_names
            ws.append(headers)

            # Apply header formatting
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border

                # Adjust column width
                ws.column_dimensions[get_column_letter(col)].width = max(15, len(header) + 2)

            # Write data rows - each document becomes one row
            for doc in documents:
                row_data = [doc.get('_source_document', 'Unknown')]

                # Add data for each field column
                for field in field_names:
                    value = doc.get(field, '')

                    # Clean and format the value
                    if isinstance(value, (dict, list)):
                        value = str(value)
                    elif value is None:
                        value = ''
                    elif isinstance(value, str):
                        # Clean up common formatting issues
                        value = value.strip().replace('\n', ' ').replace('\r', '')
                        # Remove quotes if they wrap the entire value
                        if value.startswith('"') and value.endswith('"'):
                            value = value[1:-1]

                    row_data.append(value)

                ws.append(row_data)

                # Apply cell formatting
                for col in range(1, len(row_data) + 1):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.border = self.border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Add filters and freeze panes
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = 'A2'

            # Add template info sheet
            ws_info = wb.create_sheet("Template Info")
            self._write_template_info_sheet(ws_info, documents, field_names)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return output.read()

        except Exception as e:
            logger.error(f"Error exporting template: {str(e)}")
            raise

    def _write_template_info_sheet(self, ws, documents: List[Dict[str, Any]], field_names: List[str]):
        """Write information about the template"""
        ws.append(["Document Template Information"])
        ws['A1'].font = Font(bold=True, size=16)
        ws.append([])

        # Template summary
        ws.append(["Template Summary"])
        ws['A3'].font = Font(bold=True, size=14)
        ws.append(["Total Documents", len(documents)])
        ws.append(["Total Fields Detected", len(field_names)])
        ws.append(["Generated Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])

        # Field list
        ws.append(["Detected Fields"])
        ws['A8'].font = Font(bold=True, size=14)

        for i, field in enumerate(field_names, 1):
            ws.append([f"{i}.", field])

        # Usage instructions
        ws.append([])
        ws.append(["How to Use This Template"])
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=14)

        instructions = [
            "1. Each row represents one scanned document",
            "2. Each column represents a detected field from your documents",
            "3. Data is automatically populated from AI extraction",
            "4. You can filter, sort, and analyze the data",
            "5. Use this as a base for further data processing"
        ]

        for instruction in instructions:
            ws.append([instruction])

        # Format the info sheet
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.border
                if cell.row in [1, 3, 8, ws.max_row - len(instructions)]:
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Adjust column width
        ws.column_dimensions['A'].width = 40